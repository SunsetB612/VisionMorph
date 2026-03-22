"""
结果展示服务
"""
import os
import re
from typing import List, Dict, Any, Optional

from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import load_workbook

from app.core.config import settings
from app.modules.result.schemas import (
    ResultListResponse, 
    ResultDetailResponse, 
    ResultImageInfo, 
    ResultDetailInfo,
    StaticImageResult,
    StaticResultResponse
)

OUTPUT_BASE_DIR = os.path.join(settings.BASE_DIR, "output")
EXCEL_FILENAME = "构图分析报告.xlsx"
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp")


def _sort_numeric_key(value: str) -> tuple[int, str]:
    """用于对目录名称进行数字优先排序"""
    try:
        return (0, f"{int(value):04d}")
    except ValueError:
        return (1, value.lower())


def _list_available_input_keys() -> List[str]:
    """列出可用的输入示例目录"""
    if not os.path.isdir(OUTPUT_BASE_DIR):
        return []
    keys = [
        entry for entry in os.listdir(OUTPUT_BASE_DIR)
        if os.path.isdir(os.path.join(OUTPUT_BASE_DIR, entry))
    ]
    keys.sort(key=_sort_numeric_key)
    return keys


def _resolve_input_key(requested_key: Optional[str]) -> str:
    """根据用户请求解析输入示例目录"""
    available_keys = _list_available_input_keys()
    if not available_keys:
        raise ValueError("未找到任何可用的输出目录")
    if requested_key and requested_key in available_keys:
        return requested_key
    if requested_key:
        print(f"⚠️ 未找到指定的输入示例 {requested_key}，使用默认示例 {available_keys[0]}")
    return available_keys[0]


def _list_group_directories(input_key: str) -> List[str]:
    """列出指定输入示例下的所有分组目录"""
    input_dir = os.path.join(OUTPUT_BASE_DIR, input_key)
    if not os.path.isdir(input_dir):
        return []
    groups = [
        entry for entry in os.listdir(input_dir)
        if os.path.isdir(os.path.join(input_dir, entry))
    ]
    groups.sort(key=_sort_numeric_key)
    return groups


def _read_excel_metadata(excel_path: str) -> List[Dict[str, Any]]:
    """读取 Excel 元数据"""
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    metadata: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        entry: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            entry[header] = row[idx]
        metadata.append(entry)
    return metadata


def _excel_cell_str(value: Any) -> Optional[str]:
    """将 Excel 单元格转为去首尾空格的字符串，空则 None"""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _select_image_file(image_name: str, image_files: List[str], used_files: set[str]) -> Optional[str]:
    """根据图片名字匹配对应的裁剪图片文件"""
    if not image_files:
        return None
    if image_name:
        match = re.search(r"(\d+)", image_name)
        if match:
            index = int(match.group(1)) - 1
            if 0 <= index < len(image_files):
                candidate = image_files[index]
                if candidate not in used_files:
                    return candidate
    for filename in image_files:
        if filename not in used_files:
            return filename
    return image_files[0]


def _build_static_results_for_group(input_key: str, group: str) -> List[StaticImageResult]:
    group_dir = os.path.join(OUTPUT_BASE_DIR, input_key, group)
    if not os.path.isdir(group_dir):
        return []

    excel_path = os.path.join(group_dir, EXCEL_FILENAME)
    if not os.path.exists(excel_path):
        return []
    metadata_rows = _read_excel_metadata(excel_path)
    if not metadata_rows:
        return []
    image_files = sorted([
        filename for filename in os.listdir(group_dir)
        if filename.lower().endswith(IMAGE_EXTENSIONS)
    ])
    used_files: set[str] = set()
    group_results: List[StaticImageResult] = []
    for idx, row in enumerate(metadata_rows):
        image_name = str(row.get("图片名字") or f"图片{idx + 1}")
        filename = _select_image_file(image_name, image_files, used_files)
        if not filename:
            continue
        used_files.add(filename)
        raw_score = row.get("构图分数")
        try:
            score = float(raw_score) if raw_score is not None else 0.0
        except (TypeError, ValueError):
            score = 0.0
        relative_path = f"/output/{input_key}/{group}/{filename}".replace("\\", "/")
        group_results.append(
            StaticImageResult(
                id=f"{input_key}-{group}-{filename}",
                group=group,
                image_name=image_name,
                filename=filename,
                relative_path=relative_path,
                overall_score=score,
                viewpoint_feature=_excel_cell_str(row.get("一句话概括优势特征")),
                composition_highlights=_excel_cell_str(row.get("推荐视角优点")),
                operation_guide=_excel_cell_str(row.get("操作指南")),
                orientation=_excel_cell_str(row.get("方位说明")),
                crop_type=_excel_cell_str(row.get("裁剪类型")),
            )
        )
    return group_results


def get_static_output_results(input_key: Optional[str] = None) -> StaticResultResponse:
    """获取指定输入示例下的所有图片结果，按评分排序"""
    combined_results: List[StaticImageResult] = []
    resolved_input_key = _resolve_input_key(input_key)
    group_directories = _list_group_directories(resolved_input_key)
    for group in group_directories:
        combined_results.extend(_build_static_results_for_group(resolved_input_key, group))
    sorted_results = sorted(
        combined_results,
        key=lambda item: item.overall_score,
        reverse=True
    )
    return StaticResultResponse(
        total_count=len(sorted_results),
        results=sorted_results
    )


def get_results_by_original_image(db: Session, original_image_id: int) -> ResultListResponse:
    """获取原始图片对应的所有生成图片结果，按评分从高到低排序"""
    
    try:
        # 获取原始图片信息
        original_image = db.execute(text("""
            SELECT id, user_id FROM images WHERE id = :original_image_id
        """), {"original_image_id": original_image_id}).fetchone()
        
        if not original_image:
            raise ValueError("未找到原始图片")
        
        # 获取所有生成图片及其评分，按评分从高到低排序
        results = db.execute(text("""
            SELECT 
                gi.id as generated_image_id,
                gi.filename,
                gi.file_path,
                COALESCE(ie.overall_score, 0) as overall_score,
                ie.highlights,
                gi.created_at
            FROM generated_images gi
            LEFT JOIN image_evaluations ie ON gi.id = ie.generated_image_id
            WHERE gi.original_image_id = :original_image_id
            ORDER BY COALESCE(ie.overall_score, 0) DESC, gi.created_at DESC
        """), {"original_image_id": original_image_id}).fetchall()
        
        if not results:
            raise ValueError("未找到生成图片")
        
        # 构建结果列表
        result_list = []
        for row in results:
            result_list.append(ResultImageInfo(
                generated_image_id=row[0],
                filename=row[1],
                file_path=row[2],
                overall_score=row[3],
                highlights=row[4],
                created_at=row[5]
            ))
        
        return ResultListResponse(
            original_image_id=original_image_id,
            total_count=len(result_list),
            results=result_list
        )
        
    except Exception as e:
        print(f"获取结果列表失败: {e}")
        raise ValueError(f"获取结果列表失败: {str(e)}")

def get_result_detail(db: Session, generated_image_id: int) -> ResultDetailResponse:
    """获取特定生成图片的详细结果信息"""
    
    try:
        # 获取生成图片的详细信息
        result = db.execute(text("""
            SELECT 
                gi.id as generated_image_id,
                gi.filename,
                gi.file_path,
                COALESCE(ie.overall_score, 0) as overall_score,
                ie.highlights,
                ie.ai_comment,
                ie.shooting_guidance,
                gi.created_at
            FROM generated_images gi
            LEFT JOIN image_evaluations ie ON gi.id = ie.generated_image_id
            WHERE gi.id = :generated_image_id
        """), {"generated_image_id": generated_image_id}).fetchone()
        
        if not result:
            raise ValueError("未找到生成图片")
        
        # 构建详细信息
        detail_info = ResultDetailInfo(
            generated_image_id=result[0],
            filename=result[1],
            file_path=result[2],
            overall_score=result[3],
            highlights=result[4],
            ai_comment=result[5],
            shooting_guidance=result[6],
            created_at=result[7]
        )
        
        return ResultDetailResponse(result=detail_info)
        
    except Exception as e:
        print(f"获取结果详情失败: {e}")
        raise ValueError(f"获取结果详情失败: {str(e)}")

def get_user_results(db: Session, user_id: int, limit: int = 50) -> List[ResultListResponse]:
    """获取用户的所有结果，按原始图片分组"""
    
    try:
        # 获取用户的所有原始图片
        original_images = db.execute(text("""
            SELECT id FROM images 
            WHERE user_id = :user_id 
            ORDER BY created_at DESC
            LIMIT :limit
        """), {"user_id": user_id, "limit": limit}).fetchall()
        
        if not original_images:
            return []
        
        # 为每个原始图片获取结果
        user_results = []
        for original_image in original_images:
            try:
                result = get_results_by_original_image(db, original_image[0])
                if result.total_count > 0:  # 只包含有生成图片的结果
                    user_results.append(result)
            except ValueError:
                # 如果某个原始图片没有生成图片，跳过
                continue
        
        return user_results
        
    except Exception as e:
        print(f"获取用户结果失败: {e}")
        raise ValueError(f"获取用户结果失败: {str(e)}")